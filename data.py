# pentru a ne ajuta sa trecem prin sheetul de excel
from openpyxl import load_workbook

date = load_workbook("data/first_sheet_from_data.xlsx")
# ce doc de excel alegem

sheet = date.active  # alegem acel sheet din doc excel din care dorim sa extragem datele

headers = [cell.value for cell in sheet[
    # luam header-urile pentru a putea in dictionarele din lista data sa putem pune header(key) : valoare(value)
    1]]

# nu dorim primele 2 coloane deoarece acestea ne vor impiedica sa gasim duplicatele
filtered_headers = headers[2:]

data = []

# mergem prin fiecare rand pentru a-l putea face dictionat(trecem peste primul rand care are numele coloanelor)
nr = 2  # pentru a afisa linia la care apare eroarea
none_dict = {}
for row in sheet.iter_rows(min_row=2, values_only=True):
    row_dict = {}
    for header, cell in zip(filtered_headers, row[2:]):
        try:
            if cell is None:  # aici vrem sa evidentiem ca s-a observat un None si ca trebuie inlocuit cu "-"
                if header not in none_dict:
                    none_dict[header] = 1
                else:
                    none_dict[header] += 1
                raise ValueError("S-a gasit un None:(.")
            row_dict[header] = cell
        except ValueError as e:
            print(f"Eroare: {e} pentru header-ul: '{header}', la linia {
                  nr}. Acesta va primi '-' in locul None-ului.")
            row_dict[header] = "-"
    nr += 1
    data.append(row_dict)

lungime_initiala = len(data)

data_without_duplicates = []

# eliminam duplicatele
for i in data:
    try:
        if i in data_without_duplicates:
            raise ValueError("S-a gasit un duplicat.")
        data_without_duplicates.append(i)
    except ValueError as e:
        print(
            # dorim sa scoatem in evidenta duplicatele printr-o eroare care nu opreste rularea
            f"Eroare: {e} pentru: {i}. Vom elimina acest duplicat.")

print(f"\nS-au sters {len(data) - len(data_without_duplicates)} duplicate.")
print(
    f"    ->Lungimea listei inainte de a elimina duplicatele: {lungime_initiala}.")
print(
    f"    ->Lungimea listei dupa ce am eliminat duplicatele: {len(data_without_duplicates)}.")

i = 1
print("\nVom afisa fiecare coloana pe care am gasit valori de tip 'None':")
for key in none_dict:
    print(f"    {
          i}. S-au gasit casete goale pe coloana cu numele: '{key}' de {none_dict[key]} ori.")
    i += 1

# print(data_without_duplicates)
